import logging
import os
import time
from datetime import datetime
from pathlib import Path

import gspread
from gspread.exceptions import APIError

from utils.db import get_db

log = logging.getLogger(__name__)

EXCEL_PATH = Path(r"D:\More Green AGI\More_Green_Influencer_Tracker.xlsx")
COMMENTS_HEADER = "Comments"
GSHEET_TAB = "influencers tracker"

STATUS_LABELS = {
    "discovered": "To Outreach",
    "approved": "To Outreach",
    "dm_draft": "DM Ready",
    "emailed": "Contacted",
    "contacted": "Contacted",
    "replied": "Replied",
    "address_collected": "Address Received",
    "onboarded": "Confirmed",
    "declined": "Declined",
}

SHEET_HEADERS = [
    "Instagram Handle",
    "Full Name / Channel",
    "Niche",
    "Followers",
    "Eng. Rate %",
    "Email / Contact",
    "Template Used",
    "DM Draft",
    "DM Sent Date",
    "Email Sent Date",
    "Response",
    "Collab Status",
    "Product Sent?",
    "Post Live?",
    "Notes",
    "Comments",  # Bharath's column — agent reads but NEVER overwrites
]


def _sheets_call(fn, *args, retries=5, **kwargs):
    for attempt in range(retries):
        try:
            return fn(*args, **kwargs)
        except APIError as e:
            if e.response.status_code == 429 and attempt < retries - 1:
                wait = 2 ** attempt * 15
                log.warning("Sheets rate limit — retrying in %ds", wait)
                time.sleep(wait)
            else:
                raise


def run(dry_run: bool = False) -> None:
    db = get_db()
    influencers = db.execute("SELECT * FROM influencers ORDER BY created_at DESC").fetchall()

    # Step 1: DB → Excel (source of truth file)
    _sync_db_to_excel(influencers, dry_run)

    # Step 2: DB → "influencers tracker" tab in Google Sheet
    gc = gspread.service_account(
        filename=os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE", "service_account.json")
    )
    sh = gc.open_by_key(os.environ["INFLUENCER_SHEETS_ID"])
    ws = _get_or_create_tab(sh, GSHEET_TAB)
    _sync_db_to_sheet(influencers, ws, dry_run)

    sheet_url = f"https://docs.google.com/spreadsheets/d/{os.environ['INFLUENCER_SHEETS_ID']}"
    log.info("Tracker updated. View at: %s", sheet_url)


# ── Excel ─────────────────────────────────────────────────────────────────────

def _sync_db_to_excel(influencers, dry_run: bool) -> None:
    try:
        import openpyxl
    except ImportError:
        log.warning("openpyxl not installed — skipping Excel update. Run: pip install openpyxl")
        return

    # Read existing Comments so we can preserve them
    existing_comments: dict[str, str] = {}
    if EXCEL_PATH.exists():
        wb_old = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
        ws_old = wb_old.active
        old_rows = list(ws_old.iter_rows(values_only=True))
        if old_rows:
            try:
                comments_col = list(old_rows[0]).index(COMMENTS_HEADER)
                handle_col = 0
                for row in old_rows[1:]:
                    if row and row[handle_col]:
                        handle = str(row[handle_col]).lstrip("@").strip()
                        comment = row[comments_col] if comments_col < len(row) else ""
                        if comment:
                            existing_comments[handle] = str(comment)
            except ValueError:
                pass
        wb_old.close()

    if dry_run:
        log.info("[DRY RUN] Would write %d rows to Excel", len(influencers))
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Influencers"
    ws.append(SHEET_HEADERS)

    for inf in influencers:
        row = _build_row(inf, SHEET_HEADERS)
        handle = inf["handle"]
        # Restore preserved comment
        if handle in existing_comments:
            comments_idx = SHEET_HEADERS.index(COMMENTS_HEADER)
            row[comments_idx] = existing_comments[handle]
        ws.append(row)

    wb.save(str(EXCEL_PATH))
    log.info("Excel updated: %d rows written to %s", len(influencers), EXCEL_PATH)


# ── Google Sheet ──────────────────────────────────────────────────────────────

def _get_or_create_tab(sh, title: str):
    try:
        return sh.worksheet(title)
    except gspread.exceptions.WorksheetNotFound:
        log.info("Creating new tab '%s'", title)
        return sh.add_worksheet(title=title, rows=500, cols=len(SHEET_HEADERS))


def _sync_db_to_sheet(influencers, ws, dry_run: bool) -> None:
    # Preserve existing Comments from the sheet before clearing
    existing_comments: dict[str, str] = {}
    existing_rows = ws.get_all_values()
    if len(existing_rows) > 1 and existing_rows[0]:
        header = existing_rows[0]
        if COMMENTS_HEADER in header:
            ci = header.index(COMMENTS_HEADER)
            for row in existing_rows[1:]:
                if row and row[0]:
                    handle = row[0].lstrip("@").strip()
                    comment = row[ci] if ci < len(row) else ""
                    if comment:
                        existing_comments[handle] = comment

    if dry_run:
        log.info("[DRY RUN] Would write %d rows to Sheet tab '%s'", len(influencers), GSHEET_TAB)
        return

    # Clear and rewrite the tab
    _sheets_call(ws.clear)
    all_rows = [SHEET_HEADERS]
    for inf in influencers:
        row = _build_row(inf, SHEET_HEADERS)
        handle = inf["handle"]
        if handle in existing_comments:
            ci = SHEET_HEADERS.index(COMMENTS_HEADER)
            row[ci] = existing_comments[handle]
        all_rows.append(row)

    _sheets_call(ws.update, all_rows, value_input_option="USER_ENTERED")
    log.info("Sheet tab '%s' updated: %d rows", GSHEET_TAB, len(influencers))


# ── Helpers ───────────────────────────────────────────────────────────────────

def _build_row(inf, header_row: list) -> list:
    follower_count = inf["follower_count"]
    eng_rate = inf["engagement_rate"] or 0.0
    status_label = STATUS_LABELS.get(inf["status"] or "discovered", inf["status"] or "")

    field_map = {
        "Instagram Handle": f"@{inf['handle']}",
        "Full Name / Channel": inf["full_name"] or "",
        "Niche": _humanise_hashtag(inf["source_hashtag"] or ""),
        "Followers": follower_count or "",
        "Eng. Rate %": f"{eng_rate * 100:.1f}%" if eng_rate else "",
        "Email / Contact": inf["email"] or "",
        "Template Used": inf["template_used"] or "",
        "DM Draft": inf["dm_draft_text"] or "",
        "DM Sent Date": _fmt_date(inf["dm_draft_generated_at"]),
        "Email Sent Date": _fmt_date(inf["outreach_sent_at"]),
        "Response": (inf["last_reply_preview"] or "")[:80],
        "Collab Status": status_label,
        "Product Sent?": "Yes" if inf["product_shipped"] else "No",
        "Post Live?": inf["post_live_url"] or "",
        "Notes": inf["notes"] or "",
        "Comments": "",
    }
    return [field_map.get(h, "") for h in header_row]


def _humanise_hashtag(tag: str) -> str:
    return tag.replace("_", " ").title() if tag else ""


def _fmt_date(dt_str: str | None) -> str:
    if not dt_str:
        return ""
    try:
        dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
        return dt.strftime("%d %b %Y")
    except ValueError:
        return dt_str[:10] if dt_str else ""
